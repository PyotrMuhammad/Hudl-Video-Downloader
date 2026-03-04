"""
Hudl Downloader Utilities
Common helpers: file naming, FFmpeg detection, formatting.
"""

import csv
import os
import re
import shutil
import subprocess
import sys


def find_ffmpeg() -> str:
    """Find FFmpeg binary. Returns path or raises error."""
    # Check if bundled with the app (PyInstaller)
    if getattr(sys, 'frozen', False):
        base = sys._MEIPASS
        ffmpeg = os.path.join(base, "ffmpeg.exe")
        if os.path.isfile(ffmpeg):
            return ffmpeg

    # Check PATH
    ffmpeg = shutil.which("ffmpeg")
    if ffmpeg:
        return ffmpeg

    # Common Windows locations
    common_paths = [
        r"C:\ffmpeg\bin\ffmpeg.exe",
        r"C:\Program Files\ffmpeg\bin\ffmpeg.exe",
        os.path.expanduser(r"~\ffmpeg\bin\ffmpeg.exe"),
    ]
    for p in common_paths:
        if os.path.isfile(p):
            return p

    raise FileNotFoundError(
        "FFmpeg not found! Install it:\n"
        "  Windows: https://www.gyan.dev/ffmpeg/builds/ (add to PATH)\n"
        "  Mac: brew install ffmpeg\n"
        "  Linux: sudo apt install ffmpeg"
    )


def sanitize_filename(name: str) -> str:
    """Make a string safe for use as a filename."""
    name = re.sub(r'[<>:"/\\|?*]', '_', name)
    name = re.sub(r'\s+', ' ', name).strip()
    name = name.strip('. ')
    return name[:150] if name else "hudl_video"


def format_size(bytes_val: int) -> str:
    """Format bytes as human-readable string."""
    if bytes_val < 1024:
        return f"{bytes_val} B"
    elif bytes_val < 1024 * 1024:
        return f"{bytes_val / 1024:.1f} KB"
    elif bytes_val < 1024 * 1024 * 1024:
        return f"{bytes_val / (1024 * 1024):.1f} MB"
    else:
        return f"{bytes_val / (1024 * 1024 * 1024):.2f} GB"


def format_duration(seconds: float) -> str:
    """Format seconds as HH:MM:SS or MM:SS."""
    seconds = int(seconds)
    if seconds < 0:
        return "??:??"
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    if h > 0:
        return f"{h}:{m:02d}:{s:02d}"
    return f"{m}:{s:02d}"


def format_speed(bytes_per_sec: float) -> str:
    """Format download speed."""
    if bytes_per_sec <= 0:
        return "-- MB/s"
    if bytes_per_sec < 1024 * 1024:
        return f"{bytes_per_sec / 1024:.0f} KB/s"
    return f"{bytes_per_sec / (1024 * 1024):.1f} MB/s"


def get_unique_filepath(directory: str, filename: str) -> str:
    """Get a unique filepath, adding (1), (2), etc. if file exists."""
    base, ext = os.path.splitext(filename)
    filepath = os.path.join(directory, filename)

    counter = 1
    while os.path.exists(filepath):
        filepath = os.path.join(directory, f"{base} ({counter}){ext}")
        counter += 1

    return filepath


def read_urls_from_file(filepath: str) -> list:
    """
    Read URLs from a file. Supports .txt, .csv, and .xlsx files.
    Scans all cells/lines for anything that looks like a URL (http/hudl/m3u8).
    Returns a list of URL strings.
    """
    filepath = os.path.abspath(filepath)
    ext = os.path.splitext(filepath)[1].lower()

    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")

    urls = []

    if ext == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for .xlsx files: pip install openpyxl")
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        found = _extract_url_from_text(str(cell))
                        if found:
                            urls.append(found)
        wb.close()

    elif ext == ".csv":
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                for cell in row:
                    if cell:
                        found = _extract_url_from_text(cell.strip())
                        if found:
                            urls.append(found)

    else:
        # .txt or any other text file — one URL per line
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    found = _extract_url_from_text(line)
                    if found:
                        urls.append(found)

    # Deduplicate while preserving order
    seen = set()
    unique = []
    for url in urls:
        if url not in seen:
            seen.add(url)
            unique.append(url)

    return unique


def _extract_url_from_text(text: str) -> str:
    """Extract a Hudl/m3u8 URL from a text string. Returns URL or empty string."""
    text = text.strip()
    # If the whole string is a URL
    if text.startswith("http://") or text.startswith("https://"):
        if "hudl" in text or "m3u8" in text or "blueframe" in text:
            return text
    # Try to find a URL inside the text
    match = re.search(r'(https?://[^\s<>"\']+(?:hudl|m3u8|blueframe)[^\s<>"\']*)', text)
    if match:
        return match.group(1)
    return ""


def get_ffmpeg_version(ffmpeg_path: str) -> str:
    """Get FFmpeg version string."""
    try:
        result = subprocess.run(
            [ffmpeg_path, "-version"],
            capture_output=True, text=True, timeout=5
        )
        first_line = result.stdout.split("\n")[0]
        return first_line
    except Exception:
        return "unknown"
